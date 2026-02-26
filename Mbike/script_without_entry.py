from __future__ import annotations
from rdflib import Graph
from string import Template
from InquirerPy import inquirer
from graphviz import Digraph
from openpyxl import Workbook
import subprocess

from gemseo.utils.study_analyses.coupling_study_analysis import CouplingStudyAnalysis

from nlp import requirement_extraction
################################################################################################################### 
#                       Displaying of maintenance requirements applied to an Mbike subsytem                       #
###################################################################################################################

#############  Load of the graph #############  
g = Graph()
g.parse("data/save 2/bikeo.ttl", format="turtle") 

############# Generic functions #############  

def drop_down_list(message, results_query, options=[], default=None, tuple_choice=0):
    """ Print a drop-down list from the first tuple of a SPARQL query """
    if results_query :
        options = []
        for row in results_query:
            elt = str(row[tuple_choice]) if tuple_choice != 0 else str(row[0])    #row[0] match the value of the first tuple for each column of a sparql query            
            options.append(elt)  
    choice = inquirer.select(
        message=message,
        choices=options,
        default=default
    ).execute()
    return choice

def run_query(query_template, **params):
    """ Run a template-based SPARQL query """
    query = Template(query_template).substitute(params)
    return g.query(query)

def print_columns(query_results):
    """ Print value results for each columns from a SPARQL query  """
    for column in query_results:
        for label, value in column.labels.items():
            if value == 0:
                print(str("\n" + column[label]), ": ", end='')
            elif value > 0 and value < (len(column.labels) - 1):
                print(str(column[label]), "- ", end='')
            else:
                print(str(column[label]))

def extract_columns(query_results): #not used
    """ Extract all columns from a SPARQL query """
    list_columns = []
    for column in query_results:
        column_dict = {}
        for label in column.labels:
            column_dict[str(label)] = str(column[label])
        list_columns.append(column_dict) 
    return list_columns

def requirement_to_constraint(requirement: Requirement):
    variable = requirement.variable.lower().replace(" ", "_")
    operateur = requirement.operateur
    valeur = requirement.valeur
    unite = requirement.unite
    constraint = f"{variable} {operateur} {valeur} {unite}"
    constraint_truncated = variable
    return constraint, constraint_truncated

############# User selection of the Mbike subsystem ############# 

query_subsystems = ("""
PREFIX bikeo: <http://purl.org/ontology/bikeo#>
PREFIX rdf: <http://www.w3.org/1999/02/22-rdf-syntax-ns#>
PREFIX rdfs: <http://www.w3.org/2000/01/rdf-schema#>

SELECT ?commonPartLabel ?commonPartComment
WHERE {
    ?commonPart rdfs:subClassOf+ bikeo:CommonPart ;   
                rdfs:label ?commonPartLabel ;
                rdfs:comment ?commonPartComment .
}
""")

results_query_subsystems = g.query(query_subsystems)
subsystem_selected = drop_down_list("Select the mbike's subsystem whose requirements you would like to view  :", results_query_subsystems, [], "Battery")

############# Displaying all requirements applied to an Mbike subsystem ############# 

query_requirements = ("""
PREFIX bikeo: <http://purl.org/ontology/bikeo#>
PREFIX rdf: <http://www.w3.org/1999/02/22-rdf-syntax-ns#>
PREFIX rdfs: <http://www.w3.org/2000/01/rdf-schema#>
SELECT ?requirementLabel ?requirementComment
WHERE {
    ?commonPartSelected rdfs:subClassOf+ bikeo:CommonPart ;   
                rdfs:label "$subsystemLabel"@en ;
                rdfs:subClassOf ?restriction .
 
    ?restriction a owl:Restriction ;
                owl:hasValue ?requirement .

    ?requirement a bikeo:Requirement ;
                rdfs:label ?requirementLabel .             

    OPTIONAL { ?requirement rdfs:comment ?requirementComment . }
}                                               
""")

results_query_requirements = run_query(query_requirements, subsystemLabel=subsystem_selected)
print("\n" + f"Requirement(s) applied on the {subsystem_selected} :")
print_columns(results_query_requirements)

############# Filtering of requirements by standards ############# 

query_standards = ("""
PREFIX bikeo: <http://purl.org/ontology/bikeo#>
PREFIX rdf: <http://www.w3.org/1999/02/22-rdf-syntax-ns#>
PREFIX rdfs: <http://www.w3.org/2000/01/rdf-schema#>
SELECT ?standardName
WHERE {
    ?commonPartSelected rdfs:subClassOf+ bikeo:CommonPart ;   
                rdfs:label "$subsystemLabel"@en ;
                rdfs:subClassOf ?restriction .
 
    ?restriction a owl:Restriction ;
                owl:hasValue ?requirement .

    ?requirement a bikeo:Requirement ;
                bikeo:isDerived ?standard .

    ?standard a bikeo:Standard ;
                rdfs:label ?standardName .
}                                               
""")

results_query_standards = run_query(query_standards, subsystemLabel=subsystem_selected)
standard_selected = drop_down_list("\n" + f"To filter {subsystem_selected} requirements select a standard from which requirements are derived : ", results_query_standards, [], "4210-2 :2023")

query_requirement_filtered = ("""
PREFIX bikeo: <http://purl.org/ontology/bikeo#>
PREFIX rdf: <http://www.w3.org/1999/02/22-rdf-syntax-ns#>
PREFIX rdfs: <http://www.w3.org/2000/01/rdf-schema#>
SELECT ?requirementLabel ?requirementComment
WHERE {
    ?commonPartSelected rdfs:subClassOf+ bikeo:CommonPart ;   
                rdfs:label "$subsystemLabel"@en ;
                rdfs:subClassOf ?restriction .
 
    ?restriction a owl:Restriction ;
                owl:hasValue ?requirement .

    ?standardSelected a bikeo:Standard ;
                rdfs:label "$standardLabel"@en .                    

    ?requirement a bikeo:Requirement ; 
                bikeo:isDerived ?standardSelected ;     
                rdfs:label ?requirementLabel .

    OPTIONAL { ?requirement rdfs:comment ?requirementComment . }
}                                               
""")

results_query_requirements_filtered = run_query(query_requirement_filtered, subsystemLabel=subsystem_selected, standardLabel=standard_selected)
requirement_selected = drop_down_list("\n" + f"From the following requirements derived from {standard_selected}, choose one to adapt it into a constraint:", results_query_requirements_filtered, [], "", 1)

################################################################################################################### 
#               Maintainability requirements adaptation into constraints, variables or disciplines                #
###################################################################################################################

result_extraction = requirement_extraction(requirement_selected)

constraint_suggested, constraint_truncated = requirement_to_constraint(result_extraction)

constraint_choice = drop_down_list("\n" + f"Would you like to incorporate the following constraint \"{constraint_suggested}\" into an optimisation matrix?", None, ["Yes", "No"], "", None)

################################################################################################################### 
#             Displaying of defined design variables similar to the one the user is seeking to define             #
###################################################################################################################

def ask_text(message, default=""):
    return inquirer.text(
        message=message,
        default=default
    ).execute()

def input_matrix(question, iteration_message, ):
    entries = []
    nb_inputs = int(ask_text(question))
    for _ in range(nb_inputs):
        entries.append(ask_text(iteration_message))
    return entries

def input_matrix_scenario():
    nb_design_variables = int(ask_text("How many design variables ?"))
    design_variables = []
    for _ in range(nb_design_variables):
        design_variable_selected = drop_down_list(message="\nWhich variable is a design variable ?:", choices=all_variables, default=None)
        design_variables.append(design_variable_selected)
 
#subsystem_selected = drop_down_list("Select the mbike's subsystem to which the XDSM matrix will apply :", results_query_subsystems, "Battery")

print("Please specify the number of disciplines, their names and the input and output variables in each discipline")

nb_disciplines = int(ask_text("How many disciplines ?"))
disciplines_info = {}
all_variables = []

for i in range(nb_disciplines):

    print("\n---------------------------------------")
    print(f"Definition of discipline #{i+1}")
    print("---------------------------------------")

    discipline_name = ask_text("Discipline name :")
    disciplines_info[discipline_name] = {}

    # Input variables
    print(f"\nInput variables of {discipline_name}")
    inputs = input_matrix(f"How many inputs variables for {discipline_name} ?", "Input variable name : ")
    disciplines_info[discipline_name]["inputs"] = inputs
    all_variables += inputs

    # Output variables
    print(f"\nOutput variables of {discipline_name}")
    outputs = input_matrix(f"How many outputs variables for {discipline_name} ?", "Output variable name : ")
    disciplines_info[discipline_name]["outputs"] = outputs  
    all_variables += outputs

all_variables = set(all_variables)
scenario = {}

# Constraints
nb_constraints_suggested = 0
nb_constraints = 0

if constraint_choice == "Yes" :
    constraints = [constraint_truncated]
    nb_constraints_suggested = 1
    constraint_discipline = inquirer.select(message=f"\nOn which discipline is applied the suggested constraint \"{constraint_suggested}\"?:", choices=disciplines_info.keys(), default=None).execute()
    disciplines_info[constraint_discipline]["outputs"].append(constraint_truncated) 
    all_variables.add(constraint_truncated) 
    nb_constraints = int(ask_text(f"How many constraints in addition to {constraint_suggested} ?"))
else :
    constraints = []
    nb_constraints = int(ask_text("How many constraints?"))

for _ in range(nb_constraints):
    constraint_selected = inquirer.select(message="\nOn which variables is there a constraint ?:", choices=all_variables, default=None).execute()
    constraints.append(constraint_selected) 

nb_constraints += nb_constraints_suggested

scenario["constraints"] = constraints

# Design Variables
nb_design_variables = int(ask_text("How many design variables ?"))
design_variables = []
for _ in range(nb_design_variables):
    design_variable_selected = inquirer.select(message="\nWhich variable is a design variable ?:", choices=all_variables, default=None).execute()
    design_variables.append(design_variable_selected)
scenario["design_variables"] = design_variables

# Objective function
objective_function = inquirer.select(message="\nWhich variable represents the objective function ? :", choices=all_variables, default=None).execute()
scenario["objective_function"] = objective_function

# Formulation 
formulation = ask_text("What is the formulation ?")
scenario["formulation"] = formulation

# Disciplines
scenario["disciplines"] = list(disciplines_info)

############# The user selects one variable to define from among all those created #############

variable_selected = inquirer.select(message="\nWhich variable do you want to define ?:", choices=all_variables, default=None).execute()

print(f"\nSelected variable: {variable_selected}\n")

############# Entering the unit of a variable #############

unit = ask_text(f"In what unit should the variable {variable_selected} be defined ?")

############# Suggestion for similar variables #############

system_type = ask_text("Specify the system type : ")
environment = ask_text("Specify the environment : ")

print(f"From optimisation problems on {system_type} with a {subsystem_selected} "
      f"and used in a {environment} environment:\n")

################################################################################################################### 
#                                              Build XDSM Matrix in an XLSX file                                  #
###################################################################################################################

def generate_excel(disciplines_info, scenario_info, output_path):
    wb = Workbook()

    # -----------------------------------
    # 1) Sheets for each discipline
    # -----------------------------------
    for i, (disc_name, disc_data) in enumerate(disciplines_info.items()):
        ws = wb.active if i == 0 else wb.create_sheet(disc_name)
        ws.title = disc_name

        # Writing headers
        ws.cell(row=1, column=1).value = "Inputs"
        ws.cell(row=1, column=3).value = "Outputs"

        # Writing values for each headers
        max_len = max(len(disc_data["inputs"]), len(disc_data["outputs"]))
        for row in range(max_len):
            ws.cell(row=row+2, column=1).value = disc_data["inputs"][row] if row < len(disc_data["inputs"]) else ""
            ws.cell(row=row+2, column=3).value = disc_data["outputs"][row] if row < len(disc_data["outputs"]) else ""

    # -----------------------------------
    # 2) Scenario sheet
    # -----------------------------------
    ws = wb.create_sheet("Scenario")

    # Wirting headers
    scenario_headers = [
        "",
        "Design variables",
        "",
        "Objective function",
        "",
        "Constraints",
        "",
        "Formulation",
        "",
        "Disciplines",
    ]

    for col, header in enumerate(scenario_headers, start=1):
        ws.cell(row=1, column=col).value = header

    # Witing values for each headers
    ws.cell(row=2, column=2).value = ", ".join(scenario_info["design_variables"])
    ws.cell(row=2, column=4).value = scenario_info["objective_function"]
    for row in range(nb_constraints) :
        ws.cell(row=row+2, column=6).value = scenario_info["constraints"][row]
    ws.cell(row=2, column=8).value = scenario_info["formulation"]
    disciplines_len = len(scenario_info["disciplines"])
    for row in range(disciplines_len) :
        ws.cell(row=row+2, column=10).value = scenario_info["disciplines"][row]
 
    # -----------------------------------
    # 3) XLSX file save
    # -----------------------------------
    wb.save(output_path)
    print(f"Fichier généré : {output_path}")

############# Build XDSM Matrix in an XLSX file #############

generate_excel(disciplines_info, scenario, "mdo_study.xlsx")
study = CouplingStudyAnalysis("mdo_study.xlsx")
subprocess.run(
    "gemseo-study mdo_study.xlsx -o ./gemse -t mdo -x --height 2 --width 5",
    shell=True,
    check=True
)