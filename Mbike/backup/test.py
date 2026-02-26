from __future__ import annotations
from openpyxl import Workbook
import subprocess

from gemseo.utils.study_analyses.coupling_study_analysis import CouplingStudyAnalysis

def generate_excel_from_dicts(disciplines_info, scenario_info, output_path):
    wb = Workbook()

    # -----------------------------------
    # 1) Feuilles pour chaque discipline
    # -----------------------------------
    for i, (disc_name, disc_data) in enumerate(disciplines_info.items()):
        ws = wb.active if i == 0 else wb.create_sheet(disc_name)
        ws.title = disc_name

        # Colonnes : Inputs | Outputs
        ws.cell(row=1, column=1).value = "Inputs"
        ws.cell(row=1, column=3).value = "Outputs"

        # Ajouter les valeurs en colonnes
        max_len = max(len(disc_data["inputs"]), len(disc_data["outputs"]))
        for row in range(max_len):
            ws.cell(row=row+2, column=1).value = disc_data["inputs"][row] if row < len(disc_data["inputs"]) else ""
            ws.cell(row=row+2, column=3).value = disc_data["outputs"][row] if row < len(disc_data["outputs"]) else ""
        max_len = max(len(disc_data["inputs"]), len(disc_data["outputs"]))
        for row in range(max_len):
            ws.cell(row=row+2, column=1).value = disc_data["inputs"][row] if row < len(disc_data["inputs"]) else ""
            ws.cell(row=row+2, column=3).value = disc_data["outputs"][row] if row < len(disc_data["outputs"]) else ""

    # -----------------------------------
    # 2) Feuille Scenario
    # -----------------------------------
    ws = wb.create_sheet("Scenario")

    # Définition des sections du scénario dans des colonnes
    scenario_headers = [
        "",
        "Design variables",
        "",
        "Objective function",
        "",
        "Constraints",
        "",
        "Formulation",
        "",
        "Disciplines",
    ]

    # Écriture des en-têtes
    for col, header in enumerate(scenario_headers, start=1):
        ws.cell(row=1, column=col).value = header

    # Ligne suivante : valeurs
    ws.cell(row=2, column=2).value = ", ".join(scenario_info["design_variables"])
    ws.cell(row=2, column=4).value = scenario_info["objective_function"]
    ws.cell(row=2, column=6).value = ", ".join(scenario_info["constraints"])
    ws.cell(row=2, column=8).value = scenario_info["formulation"]
    for row in range(len(scenario_info["disciplines"])) :
        ws.cell(row=row+2, column=10).value = scenario_info["disciplines"][row]

 
    # -----------------------------------
    # 3) Sauvegarde
    # -----------------------------------
    wb.save(output_path)
    print(f"📁 Fichier généré : {output_path}")


# -----------------------------------
# Exemple d'utilisation
# -----------------------------------

disciplines_info = {
    'a': {'inputs': ['b', 'f'], 'outputs': ['c']},
    'd': {'inputs': ['c'], 'outputs': ['e']}
}

scenario_info = {
    'design_variables': ['b'],
    'objective_function': 'c',
    'constraints': ['c'],
    'formulation': 'MDF',
    'disciplines': list(disciplines_info)
}

generate_excel_from_dicts(disciplines_info, scenario_info, "mdo_study_from_dicts.xlsx")

study = CouplingStudyAnalysis("mdo_study_from_dicts.xlsx")
subprocess.run(
    "gemseo-study mdo_study_from_dicts.xlsx -o ./gemse -t mdo -x --height 5 --width 5",
    shell=True,
    check=True
)
############# Build and display XDSM Matrix #############

# print("\n XDSM MATRIX\n")

# xdsm_matrix = build_xdsm_matrix(disciplines)

# print(tabulate(xdsm_matrix, headers=["Discipline", "Type", "Variable"], tablefmt="grid"))

# # Option: Save everything
# save_json({"subsystem": subsystem_selected, "disciplines": disciplines})


# def export_xdsm_graph(disciplines, filename="xdsm_graph", format="png"):
#     """
#     Export a graphical XDSM diagram using Graphviz.
#     disciplines structure:
#     {
#         "disc1": { "inputs": [...], "outputs": [...]},
#         "disc2": { ... }
#     }
#     """

#     g = Digraph("XDSM", filename=filename, format=format)
#     g.attr(rankdir="LR", fontsize="12")

#     # --------------------------
#     # CREATE DISCIPLINE BLOCKS
#     # --------------------------
#     for disc_name in disciplines.keys():
#         g.node(
#             disc_name,
#             label=f"<<B>{disc_name}</B>>",
#             shape="box",
#             style="filled",
#             fillcolor="#f7d794",
#             fontsize="12"
#         )

#     # --------------------------
#     # CREATE VARIABLE NODES
#     # --------------------------
#     for disc_name, info in disciplines.items():

#         # Inputs (blue ovals)
#         for input in info["inputs"]:
#             input_name = input[0]
#             # unit = input["unit"]
#             node_id = f"{disc_name}_in_{input_name}"

#             g.node(
#                 node_id,
#                 label=f"{input_name}\n[{unit}]",
#                 shape="ellipse",
#                 style="filled",
#                 fillcolor="#74b9ff",
#                 fontsize="11"
#             )

#             # Arrow from variable → discipline
#             g.edge(node_id, disc_name)

#         # Outputs (green ovals)
#         for output in info["outputs"]:
#             output_name = output[0]
#             # unit = var["unit"]
#             node_id = f"{disc_name}_out_{output_name}"

#             g.node(
#                 node_id,
#                 label=f"{output}\n[{unit}]",
#                 shape="ellipse",
#                 style="filled",
#                 fillcolor="#55efc4",
#                 fontsize="11"
#             )

#             # Arrow from discipline → variable
#             g.edge(disc_name, node_id)

#     # Render file
#     g.render(cleanup=True)
#     print(f"\nXDSM graph exported as: {filename}.{format}\n")

# export_xdsm_graph(disciplines_info, filename="xdsm_graph", format="png")